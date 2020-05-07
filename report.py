import os, json 
from datetime import date
import praw
from openpyxl import Workbook, load_workbook
import settings


class Report(object):
    def __init__(self):

        self.dirName = settings.dirName
        self.reddit = praw.Reddit(client_id=settings.c_id,
                                client_secret=settings.c_secret,
                                user_agent=settings.u_agent)
        self.post_limit = settings.post_limit


    def add_titles(self,ws):

        column_titles = ('Subredit','Post Title','Url','Score',
                'Upvote Ratio','# Comments','# Subscribers',
                'Score/Subscribers')

        c = 1
        for title in column_titles:
            ws.cell(column=c, row=1, value=title)
            c = c + 1
     
        print('Added titles to report.')


    def new_directory(self):
        
        try:
            os.mkdir(self.dirName)
            print("Directory " + self.dirName + " created ")

        except FileExistsError:
            print(settings.dirName + " directory loaded.")


    def new_report(self,report_file):

        wb = Workbook()
        print(report_file + " created.")
            
        return wb


    def generate_data(self,workbook,database):
        
        subreddits = database

        worksheet = workbook.active

        row_count = 2

        for sub in subreddits:

            for submission in self.reddit.subreddit(sub).hot(limit=self.post_limit):
                worksheet.cell(column=1, row=row_count, value=sub)
                worksheet.cell(column=2, row=row_count, value=submission.title)
                worksheet.cell(column=3, row=row_count, value=submission.url)
                worksheet.cell(column=4, row=row_count, value=submission.score)
                worksheet.cell(column=5, row=row_count, value=submission.upvote_ratio)
                worksheet.cell(column=6, row=row_count, value=submission.num_comments)
                worksheet.cell(column=7, row=row_count, value=self.reddit.subreddit(sub).subscribers)
                worksheet.cell(column=8, row=row_count, value=submission.score / self.reddit.subreddit(sub).subscribers)


                row_count = row_count + 1
            
            print(sub + ' data extracted.')

        return worksheet


    def open_report(self,report_file):

        print('Opening report.')
        os.startfile(report_file)
        print('Report Open.')


    def name_report(self, database_list, selected_db):

        current_time = str(date.today())
        report_file = settings.dirName + os.sep + 'Report_' + database_list[selected_db] + '_' + current_time + '.xlsx'
        return report_file

    def save_report(self,workbook,report_file):

        file_saved = False
        while file_saved == False:
            try:
                workbook.save(filename = report_file)
                file_saved = True
            except PermissionError:
                print('Please, close the report and to continue.')